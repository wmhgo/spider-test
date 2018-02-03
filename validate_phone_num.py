#! /usr/bin/env python3
import openpyxl
import openpyxl.worksheet.worksheet
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import sys
import re

class PhoneNumInvalidator(object):
  def __init__(self, fname):
    self.fname = fname

  def Validate(self):
    wb = openpyxl.load_workbook(self.fname)
    ws = wb.active
    ill_num = []
    for col in ws.columns:
      if (re.search(r'手机号码', col[0].value)):
        for index in range(1, len(col)):
          val = col[index].value
          if (not re.match(r'^\d{11}$', str(val))):
            ill_num.append(val)

    res = Workbook()
    for i, num in enumerate(ill_num):
      res.active.cell(column = 1, row = i + 1, value = num)
    res.save("result.xlsx")



if __name__ == "__main__":
  val = PhoneNumInvalidator(sys.argv[1])
  val.Validate()

